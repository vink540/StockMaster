from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, F, Count, Avg
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
from .models import Product, Category, StockMovement, PriceHistory
from sales.models import Sale, SaleItem, Customer
import json

@login_required
def home(request):
    """Dashboard principal mejorado"""
    company = request.user.company
    
    # Estadísticas generales
    total_products = Product.objects.filter(company=company, is_active=True).count()
    low_stock_count = Product.objects.filter(company=company, is_active=True, stock__lte=F('min_stock')).count()
    
    expiring_soon = Product.objects.filter(
        company=company,
        is_active=True,
        expiration_date__gte=timezone.now().date(),
        expiration_date__lte=timezone.now().date() + timedelta(days=30)
    ).count()
    
    # Ventas del mes
    first_day = timezone.now().replace(day=1)
    monthly_sales = Sale.objects.filter(company=company, created_at__gte=first_day)
    total_monthly_revenue = monthly_sales.aggregate(total=Sum('total'))['total'] or 0
    monthly_sales_count = monthly_sales.count()
    
    # Productos con bajo stock (top 5)
    low_stock_products = Product.objects.filter(
        company=company,
        is_active=True, 
        stock__lte=F('min_stock')
    ).order_by('stock')[:5]
    
    # Últimas ventas (top 5)
    recent_sales = Sale.objects.filter(company=company).select_related('customer').order_by('-created_at')[:5]
    
    context = {
        'total_products': total_products,
        'low_stock_count': low_stock_count,
        'expiring_soon': expiring_soon,
        'total_monthly_revenue': total_monthly_revenue,
        'monthly_sales_count': monthly_sales_count,
        'low_stock_products': low_stock_products,
        'recent_sales': recent_sales,
    }
    return render(request, 'inventory/home.html', context)


# ========== PRODUCTS ==========

@login_required
def product_list(request):
    """Lista de productos"""
    company = request.user.company
    products = Product.objects.filter(company=company, is_active=True).select_related('category')
    
    # Filtros
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    stock_filter = request.GET.get('stock', '')
    
    if search:
        products = products.filter(
            Q(name__icontains=search) | 
            Q(barcode__icontains=search) |
            Q(description__icontains=search)
        )
    
    if category_id:
        products = products.filter(category_id=category_id)
    
    if stock_filter == 'low':
        products = products.filter(stock__lte=F('min_stock'))
    elif stock_filter == 'out':
        products = products.filter(stock=0)
    
    categories = Category.objects.filter(company=company)
    
    context = {
        'products': products,
        'categories': categories,
        'search': search,
        'selected_category': category_id,
        'stock_filter': stock_filter,
    }
    return render(request, 'inventory/product_list.html', context)

@login_required
def product_detail(request, pk):
    """Detalle de producto"""
    company = request.user.company
    product = get_object_or_404(Product, pk=pk, company=company)
    stock_movements = product.stock_movements.all()[:10]
    
    context = {
        'product': product,
        'stock_movements': stock_movements,
    }
    return render(request, 'inventory/product_detail.html', context)

@login_required
def product_create(request):
    """Crear producto"""
    company = request.user.company
    
    if request.method == 'POST':
        try:
            category_id = request.POST.get('category')
            category_id = int(category_id) if category_id else None
            
            # Obtener valores (SIMPLIFICADO PARA ENTEROS)
            cost_price_str = request.POST.get('cost_price', '').strip()
            sale_price_str = request.POST.get('sale_price', '').strip()
            stock_str = request.POST.get('stock', '0').strip()
            min_stock_str = request.POST.get('min_stock', '5').strip()
            
            if not cost_price_str or not sale_price_str:
                messages.error(request, '❌ Los precios son obligatorios')
                categories = Category.objects.filter(company=company)
                return render(request, 'inventory/product_form.html', {'categories': categories})
            
            try:
                cost_price = Decimal(int(cost_price_str))
                sale_price = Decimal(int(sale_price_str))
                stock = Decimal(int(stock_str))
                min_stock = Decimal(int(min_stock_str))
            except (ValueError, TypeError):
                messages.error(request, '❌ Los valores deben ser números enteros válidos')
                categories = Category.objects.filter(company=company)
                return render(request, 'inventory/product_form.html', {'categories': categories})
            
            if cost_price < 0 or sale_price < 0 or stock < 0 or min_stock < 0:
                messages.error(request, '❌ Los valores no pueden ser negativos')
                categories = Category.objects.filter(company=company)
                return render(request, 'inventory/product_form.html', {'categories': categories})
            
            expiration_date = request.POST.get('expiration_date') or None
            
            product = Product.objects.create(
                company=company,  # ← FILTRO POR EMPRESA
                barcode=request.POST.get('barcode'),
                name=request.POST.get('name'),
                description=request.POST.get('description', ''),
                category_id=category_id,
                cost_price=cost_price,
                sale_price=sale_price,
                stock=stock,
                min_stock=min_stock,
                expiration_date=expiration_date,
                created_by=request.user
            )
            
            if request.FILES.get('image'):
                product.image = request.FILES['image']
                product.save()
            
            if product.stock > 0:
                StockMovement.objects.create(
                    product=product,
                    movement_type='IN',
                    quantity=product.stock,
                    previous_stock=0,
                    new_stock=product.stock,
                    reason='Stock inicial',
                    created_by=request.user
                )
            
            messages.success(request, f'✅ Producto "{product.name}" creado exitosamente.')
            return redirect('inventory:product_list')
        except Exception as e:
            messages.error(request, f'❌ Error al crear producto: {str(e)}')
    
    categories = Category.objects.filter(company=company)
    context = {'categories': categories}
    return render(request, 'inventory/product_form.html', context)

@login_required
def product_update(request, pk):
    """Actualizar producto - registra cambios de precio automáticamente"""
    company = request.user.company
    product = get_object_or_404(Product, pk=pk, company=company)
    
    if request.method == 'POST':
        try:
            # Obtener category_id
            category_id = request.POST.get('category')
            category_id = int(category_id) if category_id else None
            
            # Guardar precios anteriores
            old_cost_price = product.cost_price
            old_sale_price = product.sale_price
            
            # Obtener valores de precios (SIMPLIFICADO PARA ENTEROS)
            cost_price_str = request.POST.get('cost_price', '').strip()
            sale_price_str = request.POST.get('sale_price', '').strip()
            min_stock_str = request.POST.get('min_stock', '5').strip()
            
            # Validar que no estén vacíos
            if not cost_price_str or not sale_price_str:
                messages.error(request, '❌ El precio de costo y venta son obligatorios')
                categories = Category.objects.filter(company=company)
                return render(request, 'inventory/product_form.html', {
                    'product': product,
                    'categories': categories,
                    'is_update': True
                })
            
            # Convertir a enteros y luego a Decimal
            try:
                cost_price = Decimal(int(cost_price_str))
                sale_price = Decimal(int(sale_price_str))
                min_stock = Decimal(int(min_stock_str))
            except (ValueError, TypeError) as e:
                messages.error(request, '❌ Los precios deben ser números enteros válidos (Ejemplo: 1000, 500, 2500)')
                categories = Category.objects.filter(company=company)
                return render(request, 'inventory/product_form.html', {
                    'product': product,
                    'categories': categories,
                    'is_update': True
                })
            
            # Validar que sean positivos
            if cost_price < 0 or sale_price < 0 or min_stock < 0:
                messages.error(request, '❌ Los valores no pueden ser negativos')
                categories = Category.objects.filter(company=company)
                return render(request, 'inventory/product_form.html', {
                    'product': product,
                    'categories': categories,
                    'is_update': True
                })
            
            # Actualizar producto
            product.barcode = request.POST.get('barcode')
            product.name = request.POST.get('name')
            product.description = request.POST.get('description', '')
            product.category_id = category_id
            product.cost_price = cost_price
            product.sale_price = sale_price
            product.min_stock = min_stock
            
            # Fecha de vencimiento
            expiration_date_str = request.POST.get('expiration_date')
            product.expiration_date = expiration_date_str if expiration_date_str else None
            
            # Imagen
            if request.FILES.get('image'):
                product.image = request.FILES['image']
            
            product.save()
            
            # Registrar cambio de precio si hubo cambios
            if old_cost_price != product.cost_price or old_sale_price != product.sale_price:
                reason = request.POST.get('price_change_reason', '').strip() or 'Actualización de producto'
                
                PriceHistory.create_price_change(
                    product=product,
                    old_cost=old_cost_price,
                    old_sale=old_sale_price,
                    new_cost=product.cost_price,
                    new_sale=product.sale_price,
                    user=request.user,
                    reason=reason
                )
                messages.success(request, f'✅ Producto "{product.name}" actualizado y cambio de precio registrado.')
            else:
                messages.success(request, f'✅ Producto "{product.name}" actualizado exitosamente.')
            
            return redirect('inventory:product_detail', pk=product.pk)
            
        except Exception as e:
            messages.error(request, f'❌ Error al actualizar producto: {str(e)}')
            categories = Category.objects.filter(company=company)
            return render(request, 'inventory/product_form.html', {
                'product': product,
                'categories': categories,
                'is_update': True
            })
    
    categories = Category.objects.filter(company=company)
    context = {
        'product': product,
        'categories': categories,
        'is_update': True
    }
    return render(request, 'inventory/product_form.html', context)

@login_required
def product_delete(request, pk):
    """Eliminar producto (soft delete)"""
    company = request.user.company
    product = get_object_or_404(Product, pk=pk, company=company)
    
    if request.method == 'POST':
        product.is_active = False
        product.save()
        messages.success(request, f'Producto "{product.name}" eliminado.')
        return redirect('inventory:product_list')
    
    context = {'product': product}
    return render(request, 'inventory/product_confirm_delete.html', context)

# ========== CATEGORIES ==========

@login_required
def category_list(request):
    """Lista de categorías"""
    company = request.user.company
    categories = Category.objects.filter(company=company)
    context = {'categories': categories}
    return render(request, 'inventory/category_list.html', context)

@login_required
def category_create(request):
    """Crear categoría"""
    company = request.user.company
    
    if request.method == 'POST':
        try:
            category = Category.objects.create(
                company=company,  # ← FILTRO POR EMPRESA
                name=request.POST.get('name'),
                description=request.POST.get('description', '')
            )
            messages.success(request, f'Categoría "{category.name}" creada exitosamente.')
            return redirect('inventory:category_list')
        except Exception as e:
            messages.error(request, f'Error al crear categoría: {str(e)}')
    
    return render(request, 'inventory/category_form.html')

@login_required
def category_update(request, pk):
    """Actualizar categoría"""
    company = request.user.company
    category = get_object_or_404(Category, pk=pk, company=company)
    
    if request.method == 'POST':
        try:
            category.name = request.POST.get('name')
            category.description = request.POST.get('description', '')
            category.save()
            messages.success(request, f'Categoría "{category.name}" actualizada.')
            return redirect('inventory:category_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar categoría: {str(e)}')
    
    context = {'category': category, 'is_update': True}
    return render(request, 'inventory/category_form.html', context)

@login_required
def category_delete(request, pk):
    """Eliminar categoría"""
    company = request.user.company
    category = get_object_or_404(Category, pk=pk, company=company)
    
    if request.method == 'POST':
        category.delete()
        messages.success(request, f'Categoría "{category.name}" eliminada.')
        return redirect('inventory:category_list')
    
    context = {'category': category}
    return render(request, 'inventory/category_confirm_delete.html', context)

# ========== ALERTS ==========

from accounts.models import AlertSnapshot

@login_required
def alerts(request):
    """Vista general de alertas - marca como visto al entrar"""
    company = request.user.company
    
    # Obtener alertas actuales
    low_stock = Product.objects.filter(company=company, is_active=True, stock__lte=F('min_stock'))
    expiring_soon = Product.objects.filter(
        company=company,
        is_active=True,
        expiration_date__gte=timezone.now().date(),
        expiration_date__lte=timezone.now().date() + timedelta(days=30)
    )
    expired = Product.objects.filter(
        company=company,
        is_active=True,
        expiration_date__lt=timezone.now().date()
    )
    
    # Preparar datos del snapshot
    alert_data = {
        'low_stock_count': low_stock.count(),
        'expiring_soon_count': expiring_soon.count(),
        'expired_count': expired.count(),
        'high_debt_count': 0,
        'low_stock_ids': ','.join(map(str, low_stock.values_list('id', flat=True))),
        'expiring_soon_ids': ','.join(map(str, expiring_soon.values_list('id', flat=True))),
        'expired_ids': ','.join(map(str, expired.values_list('id', flat=True))),
        'high_debt_ids': '',
    }
    
    # Calcular clientes con deuda alta si está habilitado
    from accounts.models import SystemConfig
    
    config = SystemConfig.get_config()
    if config.enable_credits and config.enable_customers:
        high_debt_customers = []
        customers = Customer.objects.filter(company=company, is_active=True)
        for customer in customers:
            if customer.available_credit < (customer.credit_limit * Decimal('0.2')):
                high_debt_customers.append(customer.id)
        alert_data['high_debt_count'] = len(high_debt_customers)
        alert_data['high_debt_ids'] = ','.join(map(str, high_debt_customers))
    
    # Crear snapshot (marcar como visto)
    AlertSnapshot.create_snapshot(request.user, alert_data)
    
    context = {
        'low_stock': low_stock,
        'expiring_soon': expiring_soon,
        'expired': expired,
    }
    return render(request, 'inventory/alerts.html', context)

@login_required
def expiring_products(request):
    """Productos por vencer"""
    company = request.user.company
    days = int(request.GET.get('days', 30))
    products = Product.objects.filter(
        company=company,
        is_active=True,
        expiration_date__gte=timezone.now().date(),
        expiration_date__lte=timezone.now().date() + timedelta(days=days)
    )
    
    context = {'products': products, 'days': days}
    return render(request, 'inventory/expiring_products.html', context)

@login_required
def low_stock_products(request):
    """Productos con stock bajo"""
    company = request.user.company
    products = Product.objects.filter(company=company, is_active=True, stock__lte=F('min_stock'))
    context = {'products': products}
    return render(request, 'inventory/low_stock_products.html', context)

# ========== BARCODE SCANNER ==========

@login_required
def barcode_scanner(request):
    """Escáner de códigos de barras"""
    return render(request, 'inventory/barcode_scanner.html')

@login_required
def search_barcode(request):
    """API para buscar producto por código de barras"""
    company = request.user.company
    barcode = request.GET.get('barcode', '')
    
    try:
        product = Product.objects.get(company=company, barcode=barcode, is_active=True)
        data = {
            'found': True,
            'id': product.id,
            'name': product.name,
            'barcode': product.barcode,
            'price': float(product.sale_price),
            'stock': product.stock,
            'category': product.category.name if product.category else 'Sin categoría',
            'image': product.image.url if product.image else None,
        }
    except Product.DoesNotExist:
        data = {'found': False}
    
    return JsonResponse(data)

# ========== STOCK MOVEMENTS ==========

@login_required
def stock_movements(request):
    """Historial de movimientos de stock"""
    company = request.user.company
    movements = StockMovement.objects.filter(product__company=company)[:50]
    context = {'movements': movements}
    return render(request, 'inventory/stock_movements.html', context)

@login_required
def adjust_stock(request, pk):
    """Ajustar stock de producto"""
    company = request.user.company
    product = get_object_or_404(Product, pk=pk, company=company)
    
    if request.method == 'POST':
        try:
            movement_type = request.POST.get('movement_type')
            quantity = Decimal(str(request.POST.get('quantity', 0)))
            reason = request.POST.get('reason', '')
            
            previous_stock = product.stock
            
            if movement_type == 'IN':
                product.stock += quantity
            elif movement_type == 'OUT':
                product.stock -= quantity
            elif movement_type == 'ADJUSTMENT':
                product.stock = quantity
            
            product.save()
            
            StockMovement.objects.create(
                product=product,
                movement_type=movement_type,
                quantity=quantity,
                previous_stock=previous_stock,
                new_stock=product.stock,
                reason=reason,
                created_by=request.user
            )
            
            messages.success(request, 'Stock ajustado correctamente.')
            return redirect('inventory:product_detail', pk=product.pk)
        except Exception as e:
            messages.error(request, f'Error al ajustar stock: {str(e)}')
    
    context = {'product': product}
    return render(request, 'inventory/adjust_stock.html', context)

# ========== REPORTS ==========

@login_required
def reports(request):
    """Vista general de reportes"""
    return render(request, 'inventory/reports.html')

@login_required
def monthly_report(request):
    """Reporte mensual de ventas"""
    company = request.user.company
    month = request.GET.get('month', timezone.now().month)
    year = request.GET.get('year', timezone.now().year)
    
    sales = Sale.objects.filter(
        company=company,
        created_at__month=month,
        created_at__year=year
    )
    
    total_sales = sales.count()
    total_revenue = sales.aggregate(total=Sum('total'))['total'] or 0
    total_paid = sales.filter(is_paid=True).aggregate(total=Sum('total'))['total'] or 0
    total_pending = sales.filter(is_paid=False).aggregate(total=Sum('total'))['total'] or 0
    
    context = {
        'sales': sales,
        'total_sales': total_sales,
        'total_revenue': total_revenue,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'month': month,
        'year': year,
    }
    return render(request, 'inventory/monthly_report.html', context)

@login_required
def inventory_report(request):
    """Reporte de inventario"""
    company = request.user.company
    products = Product.objects.filter(company=company, is_active=True).select_related('category')
    
    total_products = products.count()
    total_value = sum(p.total_value for p in products)
    low_stock_count = products.filter(stock__lte=F('min_stock')).count()
    
    context = {
        'products': products,
        'total_products': total_products,
        'total_value': total_value,
        'low_stock_count': low_stock_count,
    }
    return render(request, 'inventory/inventory_report.html', context)

# ========== IMPORT/EXPORT EXCEL ==========

@login_required
def export_products_excel(request):
    """Exportar productos a Excel"""
    company = request.user.company
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Productos"
    
    # Encabezados
    headers = [
        'Código de Barras', 'Nombre', 'Descripción', 'Categoría',
        'Precio Costo', 'Precio Venta', 'Stock', 'Stock Mínimo',
        'Fecha Vencimiento'
    ]
    
    # Escribir encabezados con formato
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar anchos de columna
    column_widths = [18, 30, 40, 20, 15, 15, 12, 15, 18]
    for col_num, width in enumerate(column_widths, 1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = width
    
    # Obtener productos
    products = Product.objects.filter(company=company, is_active=True).select_related('category')
    
    # Escribir datos
    for row_num, product in enumerate(products, 2):
        sheet.cell(row=row_num, column=1).value = product.barcode
        sheet.cell(row=row_num, column=2).value = product.name
        sheet.cell(row=row_num, column=3).value = product.description
        sheet.cell(row=row_num, column=4).value = product.category.name if product.category else ''
        sheet.cell(row=row_num, column=5).value = float(product.cost_price)
        sheet.cell(row=row_num, column=6).value = float(product.sale_price)
        sheet.cell(row=row_num, column=7).value = float(product.stock)
        sheet.cell(row=row_num, column=8).value = float(product.min_stock)
        sheet.cell(row=row_num, column=9).value = product.expiration_date.strftime('%Y-%m-%d') if product.expiration_date else ''
        
        # Alinear números a la derecha
        for col in [5, 6, 7, 8]:
            sheet.cell(row=row_num, column=col).alignment = Alignment(horizontal='right')
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

@login_required
def import_products_excel(request):
    """Importar productos desde Excel"""
    company = request.user.company
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file:
            messages.error(request, 'Por favor seleccione un archivo.')
            return redirect('inventory:import_products_excel')
        
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser formato .xlsx')
            return redirect('inventory:import_products_excel')
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            # Saltar encabezado (fila 1)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    barcode = str(row[0]).strip() if row[0] else None
                    name = str(row[1]).strip() if row[1] else None
                    description = str(row[2]).strip() if row[2] else ''
                    category_name = str(row[3]).strip() if row[3] else None
                    cost_price = Decimal(str(row[4])) if row[4] else Decimal('0')
                    sale_price = Decimal(str(row[5])) if row[5] else Decimal('0')
                    stock = Decimal(str(row[6])) if row[6] else Decimal('0')
                    min_stock = Decimal(str(row[7])) if row[7] else Decimal('5')
                    expiration_date_str = str(row[8]).strip() if row[8] else None
                    
                    # Validar campos obligatorios
                    if not barcode or not name:
                        errors.append(f'Fila {row_num}: Faltan datos obligatorios (código o nombre)')
                        error_count += 1
                        continue
                    
                    # Procesar categoría
                    category = None
                    if category_name:
                        category, _ = Category.objects.get_or_create(
                            company=company,
                            name=category_name
                        )
                    
                    # Procesar fecha de vencimiento
                    expiration_date = None
                    if expiration_date_str:
                        try:
                            expiration_date = datetime.strptime(expiration_date_str, '%Y-%m-%d').date()
                        except:
                            pass
                    
                    # Buscar si existe el producto
                    product = Product.objects.filter(company=company, barcode=barcode).first()
                    
                    if product:
                        # Actualizar producto existente
                        product.name = name
                        product.description = description
                        product.category = category
                        product.cost_price = cost_price
                        product.sale_price = sale_price
                        product.min_stock = min_stock
                        product.expiration_date = expiration_date
                        product.save()
                        updated_count += 1
                    else:
                        # Crear nuevo producto
                        product = Product.objects.create(
                            company=company,  # ← FILTRO POR EMPRESA
                            barcode=barcode,
                            name=name,
                            description=description,
                            category=category,
                            cost_price=cost_price,
                            sale_price=sale_price,
                            stock=stock,
                            min_stock=min_stock,
                            expiration_date=expiration_date,
                            created_by=request.user
                        )
                        
                        # Registrar movimiento de stock inicial
                        if product.stock > 0:
                            StockMovement.objects.create(
                                product=product,
                                movement_type='IN',
                                quantity=product.stock,
                                previous_stock=0,
                                new_stock=product.stock,
                                reason='Importación Excel',
                                created_by=request.user
                            )
                        
                        created_count += 1
                
                except Exception as e:
                    errors.append(f'Fila {row_num}: {str(e)}')
                    error_count += 1
            
            # Mensajes de resultado
            if created_count > 0:
                messages.success(request, f'✅ {created_count} productos creados.')
            if updated_count > 0:
                messages.info(request, f'⚠️ {updated_count} productos actualizados.')
            if error_count > 0:
                messages.error(request, f'❌ {error_count} errores encontrados.')
                for error in errors[:5]:  # Mostrar solo los primeros 5 errores
                    messages.warning(request, error)
            
            return redirect('inventory:product_list')
            
        except Exception as e:
            messages.error(request, f'Error al procesar archivo: {str(e)}')
            return redirect('inventory:import_products_excel')
    
    return render(request, 'inventory/import_products_excel.html')

@login_required
def download_products_template(request):
    """Descargar plantilla de Excel para importar productos"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Productos"
    
    # Encabezados
    headers = [
        'Código de Barras', 'Nombre', 'Descripción', 'Categoría',
        'Precio Costo', 'Precio Venta', 'Stock', 'Stock Mínimo',
        'Fecha Vencimiento'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar anchos
    column_widths = [18, 30, 40, 20, 15, 15, 12, 15, 18]
    for col_num, width in enumerate(column_widths, 1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = width
    
    # Agregar ejemplos
    examples = [
        ['7791234567890', 'Coca Cola 2L', 'Gaseosa sabor cola', 'Bebidas', '500', '800', '50', '10', '2026-12-31'],
        ['7791234567891', 'Pan Lactal', 'Pan de molde integral', 'Panadería', '250', '450', '30', '5', '2025-02-15'],
        ['7791234567892', 'Arroz 1kg', 'Arroz largo fino', 'Almacén', '180', '320', '100', '20', ''],
    ]
    
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_productos.xlsx'
    
    wb.save(response)
    return response

@login_required
def price_history(request, pk):
    """Ver historial de precios de un producto específico"""
    company = request.user.company
    product = get_object_or_404(Product, pk=pk, company=company)
    history = product.price_history.all()
    
    # Preparar datos para el gráfico
    chart_data = {
        'labels': [],  # Fechas
        'cost_prices': [],  # Precios de costo
        'sale_prices': [],  # Precios de venta
        'margins': []  # Márgenes de ganancia
    }
    
    # Obtener datos históricos (últimos 20 cambios)
    for entry in reversed(list(history[:20])):
        chart_data['labels'].append(entry.changed_at.strftime('%d/%m/%Y'))
        chart_data['cost_prices'].append(float(entry.new_cost_price))
        chart_data['sale_prices'].append(float(entry.new_sale_price))
        chart_data['margins'].append(float(entry.profit_margin))
    
    # Agregar precio actual si no hay historial o es diferente al último
    if not history or (history[0].new_cost_price != product.cost_price or 
                       history[0].new_sale_price != product.sale_price):
        chart_data['labels'].append('Actual')
        chart_data['cost_prices'].append(float(product.cost_price))
        chart_data['sale_prices'].append(float(product.sale_price))
        chart_data['margins'].append(float(product.profit_margin))
    
    context = {
        'product': product,
        'history': history,
        'chart_data_json': json.dumps(chart_data),
    }
    return render(request, 'inventory/price_history.html', context)

@login_required
def all_price_changes(request):
    """Ver todos los cambios de precio recientes"""
    company = request.user.company
    changes = PriceHistory.objects.filter(
        product__company=company
    ).select_related('product', 'changed_by')[:50]
    
    context = {
        'changes': changes,
    }
    return render(request, 'inventory/all_price_changes.html', context)

@login_required
def reports_dashboard(request):
    """Dashboard principal de reportes con gráficos"""
    company = request.user.company
    
    # Período seleccionado
    period = request.GET.get('period', '30')  # días
    days = int(period)
    start_date = timezone.now() - timedelta(days=days)
    
    # Ventas del período
    sales = Sale.objects.filter(company=company, created_at__gte=start_date)
    
    # Estadísticas generales
    total_sales = sales.count()
    total_revenue = sales.aggregate(total=Sum('total'))['total'] or Decimal('0')
    avg_ticket = sales.aggregate(avg=Avg('total'))['avg'] or Decimal('0')
    
    # Productos más vendidos
    top_products = SaleItem.objects.filter(
        sale__company=company,
        sale__created_at__gte=start_date
    ).values(
        'product__name'
    ).annotate(
        total_qty=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('unit_price'))
    ).order_by('-total_revenue')[:10]
    
    # Ventas por día (últimos N días)
    daily_sales = []
    daily_labels = []
    for i in range(days-1, -1, -1):
        date = timezone.now().date() - timedelta(days=i)
        day_sales = sales.filter(created_at__date=date).aggregate(
            total=Sum('total'))['total'] or Decimal('0')
        daily_sales.append(float(day_sales))
        daily_labels.append(date.strftime('%d/%m'))
    
    # Ventas por método de pago
    payment_methods = {
        'CASH': sales.filter(payment_method='CASH').aggregate(Sum('total'))['total__sum'] or 0,
        'CARD': sales.filter(payment_method='CARD').aggregate(Sum('total'))['total__sum'] or 0,
        'TRANSFER': sales.filter(payment_method='TRANSFER').aggregate(Sum('total'))['total__sum'] or 0,
        'CREDIT': sales.filter(payment_method='CREDIT').aggregate(Sum('total'))['total__sum'] or 0,
    }
    
    context = {
        'period': period,
        'total_sales': total_sales,
        'total_revenue': total_revenue,
        'avg_ticket': avg_ticket,
        'top_products': top_products,
        'daily_sales_json': json.dumps(daily_sales),
        'daily_labels_json': json.dumps(daily_labels),
        'payment_methods_json': json.dumps({
            'labels': ['Efectivo', 'Tarjeta', 'Transferencia', 'Fiado'],
            'data': [float(payment_methods['CASH']), float(payment_methods['CARD']), 
                    float(payment_methods['TRANSFER']), float(payment_methods['CREDIT'])]
        }),
    }
    return render(request, 'reports/dashboard.html', context)

@login_required
def sales_report(request):
    """Reporte detallado de ventas"""
    company = request.user.company
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    sales = Sale.objects.filter(company=company)
    
    if start_date:
        sales = sales.filter(created_at__date__gte=start_date)
    if end_date:
        sales = sales.filter(created_at__date__lte=end_date)
    
    # Agrupar por día
    sales_by_day = sales.extra(
        select={'day': 'date(created_at)'}
    ).values('day').annotate(
        total_sales=Count('id'),
        total_revenue=Sum('total')
    ).order_by('-day')
    
    context = {
        'sales': sales,
        'sales_by_day': sales_by_day,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/sales_report.html', context)

@login_required
def products_report(request):
    """Reporte de productos más vendidos"""
    company = request.user.company
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    products_data = SaleItem.objects.filter(
        sale__company=company,
        sale__created_at__gte=start_date
    ).values(
        'product__id',
        'product__name',
        'product__barcode'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('unit_price')),
        times_sold=Count('sale', distinct=True)
    ).order_by('-total_revenue')
    
    context = {
        'products_data': products_data,
        'days': days,
    }
    return render(request, 'reports/products_report.html', context)